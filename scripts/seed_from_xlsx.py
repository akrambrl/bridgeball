"""Génère src/fleet/seed.ts à partir du classeur Excel de la flotte.
Usage: python3 scripts/seed_from_xlsx.py <fichier.xlsx>
N'utilise QUE la feuille « Véhicules » et la feuille de contraventions 2026.
Ignore les autres années et les feuilles sensibles (permis, salaires,
assurances) volontairement.
NB: read_only=False pour ne pas se fier à la dimension stockée (qui peut être
sous-estimée et tronquer les dernières lignes).
"""
import sys, json, math, re, datetime
import openpyxl

FINE_YEARS = {"2026"}

src = sys.argv[1]
wb = openpyxl.load_workbook(src, data_only=True)

def s(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()

def iso(v, fallback_year):
    if isinstance(v, (datetime.datetime, datetime.date)):
        d = v
        if d.year < 1990:
            return f"{fallback_year}-01-01"
        return f"{d.year:04d}-{d.month:02d}-{d.day:02d}"
    txt = s(v)
    m = re.match(r"^(\d{1,2})[/.-](\d{1,2})[/.-](\d{2,4})$", txt)
    if m:
        dd, mm, yy = (int(x) for x in m.groups())
        if yy < 100:
            yy += 2000
        if yy < 1990 or mm < 1 or mm > 12 or dd < 1 or dd > 31:
            return f"{fallback_year}-01-01"
        return f"{yy:04d}-{mm:02d}-{dd:02d}"
    return f"{fallback_year}-01-01"

# ---- Conducteurs (dédupliqués sur nom normalisé) ----
NON_NAMES = {"", "prénom", "prenom", "av", "conducteur", "nan", "none", "total"}
drivers = []
driver_by_key = {}

def driver_id(name):
    name = re.sub(r"\s+", " ", s(name)).strip()
    key = name.lower()
    if key in NON_NAMES:
        return None
    if key in driver_by_key:
        return driver_by_key[key]
    did = f"d{len(drivers) + 1}"
    driver_by_key[key] = did
    drivers.append({"id": did, "firstName": name, "lastName": "",
                    "email": "", "phone": "", "licenseNumber": ""})
    return did

# ---- Véhicules ----
EV = ("tesla", "marvel", "model 3", "model y", "zoe", "e-tron", "id.3", "id.4", "electr")
HYB = ("hybr", "prius")
DSL = ("diesel",)

def fuel_of(t):
    t = t.lower()
    if any(k in t for k in EV):
        return "electrique"
    if any(k in t for k in HYB):
        return "hybride"
    if any(k in t for k in DSL):
        return "diesel"
    return "essence"

vehicles = []
veh_by_driver = {}
ws = wb["Véhicules"]
rows = list(ws.iter_rows(values_only=True))
for r in rows:
    if len(r) < 3:
        continue
    conductor, plate, vtype = s(r[0]), s(r[1]), s(r[2])
    if not conductor or not plate:
        continue
    if conductor.lower() in NON_NAMES or plate.lower() in ("n° de plaque", "n de plaque"):
        continue
    if not re.search(r"\d", plate):  # une plaque contient des chiffres
        continue
    did = driver_id(conductor)
    parts = vtype.split(" ", 1) if vtype else ["Véhicule"]
    brand = parts[0] or "Véhicule"
    model = parts[1] if len(parts) > 1 else ""
    immat = r[15] if len(r) > 15 else None
    revis = r[16] if len(r) > 16 else None
    year = 0
    for d in (immat, revis):
        if isinstance(d, (datetime.datetime, datetime.date)) and d.year >= 1990:
            year = d.year
            break
    if year == 0:
        year = 2021
    km1 = r[17] if len(r) > 17 else None
    km2 = r[18] if len(r) > 18 else None
    mileage = 0
    for k in (km2, km1):
        if isinstance(k, (int, float)):
            mileage = int(k)
            break
    next_service = (math.ceil(mileage / 10000) * 10000) if mileage else 10000
    if next_service <= mileage:
        next_service = mileage + 10000
    vid = f"v{len(vehicles) + 1}"
    vehicles.append({
        "id": vid, "plate": plate, "brand": brand, "model": model, "year": year,
        "fuel": fuel_of(vtype), "status": "active", "mileage": mileage,
        "nextServiceKm": next_service, "assignedDriverId": did,
    })
    if did and did not in veh_by_driver:
        veh_by_driver[did] = vid

# ---- Entretiens déduits des dates CT / révision ----
maintenance = []
def add_maint(vid, date_cell, kind, status, mileage, notes):
    if not isinstance(date_cell, (datetime.datetime, datetime.date)):
        return
    if date_cell.year < 1990:
        return
    maintenance.append({
        "id": f"m{len(maintenance) + 1}", "vehicleId": vid,
        "date": f"{date_cell.year:04d}-{date_cell.month:02d}-{date_cell.day:02d}",
        "kind": kind, "cost": 0, "mileage": mileage, "status": status, "notes": notes,
    })

vi = 0
for r in rows:
    if len(r) < 3:
        continue
    conductor, plate = s(r[0]), s(r[1])
    if not conductor or not plate or conductor.lower() in NON_NAMES:
        continue
    if plate.lower() in ("n° de plaque", "n de plaque") or not re.search(r"\d", plate):
        continue
    v = vehicles[vi]; vi += 1
    add_maint(v["id"], r[14] if len(r) > 14 else None, "controle_technique", "done", v["mileage"], "Dernier contrôle technique")
    add_maint(v["id"], r[13] if len(r) > 13 else None, "controle_technique", "scheduled", v["mileage"], "Contrôle technique à effectuer")
    add_maint(v["id"], r[16] if len(r) > 16 else None, "revision", "done", v["mileage"], "Dernière révision")

# ---- Amendes (feuilles annuelles) ----
def status_of(cell):
    t = s(cell).lower()
    if "contest" in t:
        return "contested"
    if "attente" in t:
        return "pending"
    if "ok" in t or t in ("payé", "paye", "par tjmax !!"):
        return "paid"
    return "paid" if t else "pending"

fines = []
for name in wb.sheetnames:
    if name not in FINE_YEARS:
        continue
    year = int(name)
    wsf = wb[name]
    for r in wsf.iter_rows(values_only=True):
        if len(r) < 5:
            continue
        prenom = s(r[0])
        if not prenom or prenom.lower() in NON_NAMES:
            continue
        reason = s(r[3]) if len(r) > 3 else ""
        amount_cell = r[4] if len(r) > 4 else None
        if not reason and not isinstance(amount_cell, (int, float)):
            continue
        did = driver_id(prenom)
        amount = int(round(amount_cell)) if isinstance(amount_cell, (int, float)) else 0
        fines.append({
            "id": f"f{len(fines) + 1}",
            "vehicleId": veh_by_driver.get(did, ""),
            "driverId": did,
            "date": iso(r[2] if len(r) > 2 else None, year),
            "reason": reason or "Contravention",
            "amount": amount,
            "location": "",
            "status": status_of(r[6] if len(r) > 6 else None),
        })

fines.sort(key=lambda f: f["date"], reverse=True)

data = {"vehicles": vehicles, "drivers": drivers, "fines": fines,
        "maintenance": maintenance, "trips": []}

def arr(items):
    if not items:
        return "[]"
    return "[\n" + ",\n".join("    " + json.dumps(x, ensure_ascii=False) for x in items) + ",\n  ]"

out = (
    "import type { FleetData } from \"./types\";\n\n"
    "// Données importées depuis le classeur Excel de la flotte.\n"
    "// Régénérable via: python3 scripts/seed_from_xlsx.py <fichier.xlsx>\n"
    "export const seedData: FleetData = {\n"
    f"  vehicles: {arr(vehicles)},\n"
    f"  drivers: {arr(drivers)},\n"
    f"  fines: {arr(fines)},\n"
    f"  maintenance: {arr(maintenance)},\n"
    "  trips: [],\n"
    "};\n"
)
with open("src/fleet/seed.ts", "w", encoding="utf-8") as f:
    f.write(out)

print(f"vehicles={len(vehicles)} drivers={len(drivers)} fines={len(fines)} maintenance={len(maintenance)}")
